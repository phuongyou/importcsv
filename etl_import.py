
#!/usr/bin/env python3
"""
ETL Script: Viscosity Master Contracts Excel → PostgreSQL
=========================================================
Source:  Viscosity_Master_Contracts_File_-_CTA.xlsx
Target:  PostgreSQL (CTA Contract Management System v3.0)

Run order:
  1. business_units  (seed)
  2. user          (seed system user)
  3. clients         (from QB Consolidated Contacts + Agreements + SOWs)
  4. agreements      (from Agreements sheet)
  5. agreement_insurance (from Agreements sheet - insurance columns)
  6. sows            (from Statements of Work sheet)
  7. partnerships    (from Partnerships sheet)

Usage:
  pip install pandas openpyxl psycopg2-binary python-dotenv
  python etl_import.py
  python etl_import.py --dry-run        # parse only, no DB write
  python etl_import.py --sheet clients  # run only one step
"""

import os
import re
import sys
import argparse
import logging
import pandas as pd
import numpy as np
from datetime import datetime, date
from typing import Optional, Any
from openpyxl import load_workbook
import os

# ─── CONFIG ──────────────────────────────────────────────────────────────────
# Find the first Excel file in the parent directory

# EXCEL_PATH will be set by app.py when processing uploaded files
# Or can be set via environment variable EXCEL_PATH
EXCEL_PATH = os.getenv("EXCEL_PATH")

DB_CONFIG = {
    "host":     os.getenv("DB_HOST",     "localhost"),
    "port":     os.getenv("DB_PORT",     "5432"),
    "dbname":   os.getenv("DB_NAME",     "db_contract"),
    "user":     os.getenv("DB_USER",     "postgres"),
    "password": os.getenv("DB_PASSWORD", "123"),
}

SYSTEM_USER_ID = 1          # seeded ETL system user
DEFAULT_BU_MAP = {"VNA": 1, "VT": 2, "VT_VNA": 3, "OTHER": 4}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger("etl")

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def clean(val: Any) -> Optional[str]:
    """Return None for NaN/blank, strip whitespace otherwise."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    s = str(val).strip()
    return None if s in ("", "-", "nan", "NaT", "None") else s


def clean_date(val: Any) -> Optional[date]:
    """
    Parse dates safely.
    Excel sometimes stores expired dates as '00:00:00' or negative day-numbers.
    Supports formats: YYYY-MM-DD, MM/DD/YYYY, DD/MM/YYYY, DD.MM.YYYY
    """
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    s = str(val).strip()
    if s in ("", "-", "nan", "NaT", "None", "00:00:00"):
        return None
    # Already a datetime/date from pandas
    if isinstance(val, (datetime, pd.Timestamp)):
        try:
            d = pd.Timestamp(val)
            if d.year < 1900 or d.year > 2100:
                return None
            return d.date()
        except Exception:
            return None
    # Try parsing string
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            d = datetime.strptime(s[:10], fmt[:10])
            if 1900 <= d.year <= 2100:
                return d.date()
        except ValueError:
            continue
    return None


def clean_decimal(val: Any) -> Optional[float]:
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    s = str(val).replace(",", "").replace("$", "").strip()
    if s in ("", "-", "nan"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def clean_bool(val: Any) -> bool:
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return False
    return str(val).strip().lower() in ("yes", "1", "true", "x")


def rag_clean(val: Any) -> str:
    s = clean(val)
    if s in ("G", "A", "R"):
        return s
    return "G"


def map_bu(val: Any) -> Optional[int]:
    s = clean(val)
    if not s:
        return None
    s = s.upper().replace("-", "_")
    return DEFAULT_BU_MAP.get(s)


def map_contract_type_agr(val: Any) -> str:
    VALID = {"MSA", "NDA", "VAR Agreement", "Partnership", "Subcontracting",
             "CSA", "MCSA", "HIPAA", "Addendum", "Other"}
    s = clean(val) or "Other"
    return s if s in VALID else "Other"


def map_work_type_agr(val: Any) -> str:
    VALID = {"Agreement", "Amendment", "Partnership", "Purchase Order", "VAR Agreement"}
    RAW_MAP = {
        "professional services": "Agreement",
        "sales & use tax addendum": "Amendment",
        "unknown": "Agreement",
    }
    s = clean(val) or "Agreement"
    if s in VALID:
        return s
    return RAW_MAP.get(s.lower(), "Agreement")


def map_work_type_sow(val: Any) -> str:
    VALID = {"Managed Services", "Consulting/Professional Services", "Staff Augmentation",
             "Assessment", "T&M", "Fixed Fee", "Retainer", "Remote Support", "AOP Renewal"}
    RAW_MAP = {
        "professional services": "Consulting/Professional Services",
        "retainer services": "Retainer",
        "ms change order": "Consulting/Professional Services",
        "purchase order": "Consulting/Professional Services",
        "change order": "Consulting/Professional Services",
        "staff aug": "Staff Augmentation",
        "aop renewal": "AOP Renewal",
        "unknown": "Consulting/Professional Services",
    }
    s = clean(val) or "Consulting/Professional Services"
    if s in VALID:
        return s
    return RAW_MAP.get(s.lower(), "Consulting/Professional Services")


def map_sow_status(val: Any) -> str:
    VALID = {"Active", "Expired", "Terminated", "Archived"}
    RAW_MAP = {
        "unknown": "Active",
        "inactive": "Archived",
        "pending": "Active",
    }
    s = clean(val) or "Active"
    if s in VALID:
        return s
    return RAW_MAP.get(s.lower(), "Active")


def map_requires_po(val: Any) -> bool:
    s = clean(val)
    if not s:
        return False
    return s.lower() in ("yes", "1", "need po", "need po ")


def read_sheet(sheet_name: str) -> pd.DataFrame:
    log.info(f"Reading sheet: {sheet_name}")
    if not EXCEL_PATH or not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"No Excel file found. EXCEL_PATH={EXCEL_PATH}")
    df = pd.read_excel(EXCEL_PATH, sheet_name=sheet_name, header=0)
    # All key sheets have actual headers in row 0 (first data row)
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)
    # Drop fully empty rows
    df = df.dropna(how="all")
    return df


# ─── STEP 1 & 2: SEED business_units + system user ───────────────────────────
def seed_business_units(cur):
    log.info("Seeding business_units...")
    rows = [
        (1, "VNA",    "Viscosity North America", "North America entity"),
        (2, "VT",     "Viscosity Technology",    "Technology entity"),
        # (3, "VT_VNA", "Both Entities",           "Covers both VNA and VT"),
        (3, "OTHER",  "Other/Unknown",            "Catch-all"),
    ]
    
    inserted = 0
    for row in rows:
        # Check if business unit already exists
        cur.execute("SELECT business_units_id FROM business_units WHERE business_units_id = %s", (row[0],))
        if not cur.fetchone():
            cur.execute("""
                INSERT INTO business_units (business_units_id, code, name, description)
                VALUES (%s, %s, %s, %s)
            """, row)
            inserted += 1
    
    log.info(f"  ✓ {inserted} business_units seeded")
    return inserted





# ─── STEP 3: CLIENTS ─────────────────────────────────────────────────────────

def import_clients(cur) -> tuple:
    """
    Build client list from 3 sources (union, deduplicated by company_name):
      1. QB Consolidated Contacts (best source: has email, phone, addresses)
      2. Agreements sheet (may have clients not in QB)
      3. SOW sheet (same)
    Check for duplicates by company_name - skip if already exists.
    Returns: tuple(dict mapping company_name → client_id, inserted_count, skipped_count)
    """
    log.info("Importing clients...")

    # Source 1: QB Consolidated Contacts
    qb = pd.read_excel(EXCEL_PATH, sheet_name="QB Consolidated Contacts")
    qb_clients: dict[str, dict] = {}
    for _, row in qb.iterrows():
        name = clean(row.get("Customer"))
        if not name:
            continue
        bu_code = clean(row.get("Business"))
        bu_id = DEFAULT_BU_MAP.get((bu_code or "").upper())
        existing = qb_clients.get(name, {})
        # Merge: prefer non-null values from first encounter
        qb_clients[name] = {
            "company_name":        name,
            "business_unit_id":    existing.get("business_unit_id") or bu_id,
            "primary_email":       existing.get("primary_email")    or clean(row.get("QB Email")),
            "primary_phone":       existing.get("primary_phone")    or clean(row.get("QB Phone Numbers")),
            "primary_contact_name":existing.get("primary_contact_name") or clean(row.get("Name")) or clean(row.get("QB Full Name")),
            "billing_address":     existing.get("billing_address")  or clean(row.get("Billing Address")),
            "shipping_address":    existing.get("shipping_address") or clean(row.get("Shipping Address")),
        }

    # Source 2: Agreements
    agr_df = read_sheet("Agreements")
    agr_names = set(c for c in agr_df["Client"].dropna().unique() if clean(c))

    # Source 3: SOW
    sow_df = read_sheet("Statements of Work")
    sow_names = set(c for c in sow_df["Client"].dropna().unique() if clean(c))

    all_names = set(qb_clients.keys()) | agr_names | sow_names
    log.info(f"  QB: {len(qb_clients)}, Agr-only: {len(agr_names-set(qb_clients))}, "
             f"SOW-only: {len(sow_names-set(qb_clients)-agr_names)} → Total unique: {len(all_names)}")
    
    # DEBUG: Show all clients
    # log.info(f"\n  All clients found:")
    # for name in sorted(all_names):
    #     log.info(f"    - '{name}'")
    # log.info("")

    client_id_map: dict[str, int] = {}
    inserted = 0
    skipped = 0

    for name in sorted(all_names):
        # Check if client already exists
        cur.execute("SELECT client_id FROM clients WHERE company_name = %s", (name,))
        existing_client = cur.fetchone()
        
        if existing_client:
            # Client already exists, skip it
            client_id_map[name] = existing_client[0]
            skipped += 1
            log.debug(f"  ⊘ Client '{name}' already exists (ID: {existing_client[0]}) - skipped")
            continue
        
        meta = qb_clients.get(name, {})
        # Clean phone - remove "Phone: " prefix
        phone = meta.get("primary_phone")
        if phone:
            phone = re.sub(r'^Phone:\s*', '', phone).strip()[:50]

        cur.execute("""
            INSERT INTO clients
                (company_name, business_unit_id, client_type,
                 primary_contact_name, primary_email, primary_phone,
                 billing_address, shipping_address,
                 status, is_active, created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING client_id
        """, (
            name[:255],
            meta.get("business_unit_id"),
            "Client",
            (meta.get("primary_contact_name") or "")[:150] or None,
            (meta.get("primary_email") or "")[:255] or None,
            phone,
            meta.get("billing_address"),
            meta.get("shipping_address"),
            "Active",
            1,
            None,  # created_by: leave empty (NULL)
        ))
        row = cur.fetchone()
        client_id_map[name] = row[0]
        inserted += 1

    log.info(f"  ✓ {inserted} clients inserted, {skipped} skipped (duplicates)")
    return client_id_map, inserted, skipped


# ─── STEP 4 & 5: AGREEMENTS + INSURANCE ──────────────────────────────────────

def import_agreements(cur, client_map: dict[str, int]) -> tuple:
    """
    Returns: tuple(dict mapping Excel row-index → agreement_id, inserted_count, skipped_count)
    """
    log.info("Importing agreements + insurance...")
    df = read_sheet("Agreements")

    # Drop rows that are clearly header repeats or totally empty
    df = df[df["Client"].apply(lambda x: bool(clean(x)) and clean(x) != "Client")]

    inserted = 0
    skipped = 0
    agr_row_map: dict[int, int] = {}   # excel_idx → db id
    
    # Load Excel workbook for direct hyperlink extraction
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb["Agreements"]
        
        # Sheet structure: row 1 = dummy, row 2 = headers, row 3+ = data
        # Find Description column index in row 2 (header row)
        desc_col_idx = None
        headers = [cell.value for cell in ws[2]]
        # log.info(f"  Excel headers in row 2: {headers}")
        
        for idx, cell in enumerate(ws[2], 1):
            col_name = cell.value
            if col_name and col_name.strip() == "Description":
                desc_col_idx = idx
                # log.info(f"  ✓ Found Description column at index {desc_col_idx} (value: '{col_name}')")
                break
        
        if not desc_col_idx:
            log.warning(f"  ✗ Description column NOT found in row 2")
    except Exception as e:
        log.warning(f"  Could not load openpyxl workbook for hyperlinks: {e}")
        import traceback
        traceback.print_exc()
        wb = None
        desc_col_idx = None

    INS_COLS = [
        "GL Aggregate Limits", "GL per Occurrence",
        "Workers Compensation", "Employers Liability ",
        "Umbrella Aggregate", "Umbrella per Occurrence",
        "Cyber Aggregate", "Cyber per Occurrence",
        "Professional Liability (E&O)", "Property in Transit",
        "Employee Theft of Client Property", "Products/Completed Operations",
        "Personal & Advertising Injury", "Bodily Injury/Property Damage",
        "Fire Legal Liability Injury", "Medical Payments",
        "Auto Insurance Combined Single Limit (Hired)",
        "Auto Insurance Combined Single Limit (Non-owned)",
        "Auto Insurance Combined Single Limit (Owned)",
    ]

    for idx, row in df.iterrows():
        client_name = clean(row.get("Client"))
        if not client_name:
            skipped += 1
            continue

        client_id = client_map.get(client_name)
        if not client_id:
            log.warning(f"  ⚠ Agreement row {idx}: unknown client '{client_name}' — skipping")
            skipped += 1
            continue

        # Map business unit
        bu_raw = clean(row.get("Business"))
        bu_id = DEFAULT_BU_MAP.get((bu_raw or "").upper())

        # Map GL insurance flag
        gl_raw = clean(row.get("General Liability Insurance"))
        gl_val = None if not gl_raw else ("Yes" if gl_raw == "Yes" else ("No" if gl_raw == "No" else "Waived"))

        # COI
        coi_raw = clean(row.get("COI Requirement"))
        coi_val = "Yes" if coi_raw and coi_raw != "-" else "No"

        # Execution method: DocuSign col = Yes/No
        exec_method = "DocuSign" if clean_bool(row.get("DocuSign")) else None

        # Status — infer from RAG + expiration
        rag = rag_clean(row.get("R.A.G. Status"))
        exp_date = clean_date(row.get("Expiration Date"))
        status = "Expired" if rag == "R" else "Active"
        
        # Extract document_name (display text) and document_path (hyperlink URL)
        # Excel structure: row 1 = dummy, row 2 = headers, row 3+ = data
        # Pandas after reset_index: idx 0 = Excel row 3, so excel_row = idx + 3
        document_name = clean(row.get("Description"))
        document_path = None
        SHAREPOINT_BASE_URL = "https://viscositynorthamerica.sharepoint.com/sites/ProcessMgmt-SharedTeam/Shared Documents/General/CTA/Contracts-  Collateral/"
        
        if wb and desc_col_idx:
            try:
                excel_row = idx + 3
                cell = ws.cell(row=excel_row, column=desc_col_idx)
                if cell.hyperlink:
                    hyperlink = cell.hyperlink.target
                    # Prepend SharePoint base URL
                    document_path = SHAREPOINT_BASE_URL + hyperlink
                    # log.info(f"  Row {idx}: Excel row {excel_row} - Found hyperlink: {document_path}")
                else:
                    log.debug(f"  Row {idx}: Excel row {excel_row} - Cell value: '{cell.value}' - No hyperlink")
            except Exception as e:
                log.debug(f"  Row {idx}: Error extracting hyperlink: {e}")

        # Check if agreement already exists (by client_id + document_name + effective_date)
        eff_date = clean_date(row.get("Effective Date"))
        check_doc_name = (document_name or "")[:500] or None
        
        # Build dynamic SQL to handle NULL values
        if check_doc_name and eff_date:
            cur.execute("""
                SELECT agreements_id FROM agreements 
                WHERE client_id = %s AND document_name = %s AND effective_date = %s
            """, (client_id, check_doc_name, eff_date))
        elif check_doc_name:
            cur.execute("""
                SELECT agreements_id FROM agreements 
                WHERE client_id = %s AND document_name = %s AND effective_date IS NULL
            """, (client_id, check_doc_name))
        elif eff_date:
            cur.execute("""
                SELECT agreements_id FROM agreements 
                WHERE client_id = %s AND document_name IS NULL AND effective_date = %s
            """, (client_id, eff_date))
        else:
            cur.execute("""
                SELECT agreements_id FROM agreements 
                WHERE client_id = %s AND document_name IS NULL AND effective_date IS NULL
            """, (client_id,))
        
        if cur.fetchone():
            # Agreement already exists, skip it
            skipped += 1
            log.debug(f"  ⊘ Agreement for client_id={client_id}, doc='{document_name}' already exists - skipped")
            continue

        cur.execute("""
            INSERT INTO agreements (
                client_id, business_unit_id,
                authorized_signer, line_of_business,
                contract_type, work_type,
                document_name, document_path, document_stored, signed_copy, execution_method,
                date_signed, effective_date, duration_months, expiration_date,
                rag_status, general_liability_ins, coi_required,
                status, archived, notes,
                auto_renewal, requires_po, waiver_of_subrogation, tail_coverage_required, leadership_exemption,
                created_by, last_updated_by
            ) VALUES (
                %s,%s,  %s,%s,  %s,%s,
                %s,%s,%s,%s,%s,
                %s,%s,%s,%s,
                %s,%s,%s,
                %s,%s,%s,
                %s,%s,%s,%s,%s,
                %s,%s
            )
            RETURNING agreements_id
        """, (
            client_id, bu_id,
            (clean(row.get("Contact")) or "")[:150] or None,
            (clean(row.get("Line of Business ")) or "")[:100] or None,
            map_contract_type_agr(row.get("Contract Type")),
            map_work_type_agr(row.get("Work Type")),
            # document
            (document_name or "")[:500] or None,
            (document_path or "")[:500] or None,
            None,   # document_stored: bytea - stored as NULL initially
            ("Yes" if clean(row.get("Signed Copy")) == "Yes" else "No"),  # signed_copy: text
            exec_method,
            # dates
            clean_date(row.get("Date Signed")),
            clean_date(row.get("Effective Date")),
            (clean(row.get("Duration (Months)")) or "")[:20] or None,
            exp_date,
            # rag
            rag,
            gl_val,
            coi_val,
            # status
            status,
            1 if status == "Expired" else 0,
            (clean(row.get("Notes")) or "")[:2000] or None,
            False, False, False, False, False,  # auto_renewal, requires_po, waiver_of_subrogation, tail_coverage_required, leadership_exemption
            None, None,  # created_by, last_updated_by: leave empty
        ))
        agr_id = cur.fetchone()[0]
        agr_row_map[idx] = agr_id
        inserted += 1

        # ── Insurance record ───────────────────────────────────────────────
        # Only insert if at least one insurance field is non-null
        has_insurance = any(clean(row.get(c)) not in (None, "-") for c in INS_COLS)
        gl_required = gl_raw == "Yes"

        if has_insurance or gl_required:
            def ic(col): return (clean(row.get(col)) or "")[:50] or None
            coi_on_file = coi_val == "Yes"
            cur.execute("""
                INSERT INTO agreement_insurance (
                    agreement_id, client_id, gl_required,
                    gl_aggregate_limits, gl_per_occurrence,
                    workers_compensation, employers_liability,
                    umbrella_aggregate, umbrella_per_occurrence,
                    cyber_aggregate, cyber_per_occurrence,
                    professional_liability_eo, property_in_transit,
                    employee_theft_client_property, products_completed_operations,
                    personal_advertising_injury, bodily_injury_property_damage,
                    fire_legal_liability, medical_payments,
                    auto_csl_hired, auto_csl_non_owned, auto_csl_owned,
                    coi_on_file
                ) VALUES (
                    %s,%s,%s,
                    %s,%s, %s,%s, %s,%s, %s,%s,
                    %s,%s, %s,%s, %s,%s, %s,%s,
                    %s,%s,%s, %s
                )
                ON CONFLICT (agreement_id) DO NOTHING
            """, (
                agr_id, client_id, gl_required,
                ic("GL Aggregate Limits"), ic("GL per Occurrence"),
                ic("Workers Compensation"), ic("Employers Liability "),
                ic("Umbrella Aggregate"), ic("Umbrella per Occurrence"),
                ic("Cyber Aggregate"), ic("Cyber per Occurrence"),
                ic("Professional Liability (E&O)"), ic("Property in Transit"),
                ic("Employee Theft of Client Property"), ic("Products/Completed Operations"),
                ic("Personal & Advertising Injury"), ic("Bodily Injury/Property Damage"),
                ic("Fire Legal Liability Injury"), ic("Medical Payments"),
                ic("Auto Insurance Combined Single Limit (Hired)"),
                ic("Auto Insurance Combined Single Limit (Non-owned)"),
                ic("Auto Insurance Combined Single Limit (Owned)"),
                coi_on_file,
            ))

    if wb:
        wb.close()

    log.info(f"  ✓ {inserted} agreements inserted, {skipped} skipped")
    return agr_row_map, inserted, skipped


# ─── STEP 6: SOWs ────────────────────────────────────────────────────────────

def import_sows(cur, client_map: dict[str, int]):
    log.info("Importing SOWs...")
    df = read_sheet("Statements of Work")
    df = df[df["Client"].apply(lambda x: bool(clean(x)) and clean(x) != "Client")]

    inserted = 0
    skipped = 0
    
    # Load Excel workbook for direct hyperlink extraction
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb["Statements of Work"]
        
        # Sheet structure: row 1 = description, row 2 = headers, row 3+ = data
        # Find "Contract Name/Document Link" column index in row 2 (header row)
        doc_link_col_idx = None
        headers = [cell.value for cell in ws[2]]
        # log.info(f"  Excel headers in row 2: {headers}")
        
        for idx_col, cell in enumerate(ws[2], 1):
            col_name = cell.value
            if col_name and "Contract Name/Document Link" in str(col_name):
                doc_link_col_idx = idx_col
                log.info(f"  ✓ Found 'Contract Name/Document Link' column at index {doc_link_col_idx}")
                break
        
        if not doc_link_col_idx:
            log.warning(f"  ✗ 'Contract Name/Document Link' column NOT found in row 2")
    except Exception as e:
        log.warning(f"  Could not load openpyxl workbook for hyperlinks: {e}")
        wb = None
        doc_link_col_idx = None

    for idx, row in df.iterrows():
        client_name = clean(row.get("Client"))
        if not client_name:
            skipped += 1
            continue

        client_id = client_map.get(client_name)
        if not client_id:
            log.warning(f"  ⚠ SOW row {idx}: unknown client '{client_name}' — skipping")
            skipped += 1
            continue

        bu_raw = clean(row.get("Company"))   # Company col = VNA/VT
        bu_id = DEFAULT_BU_MAP.get((bu_raw or "").upper())

        # Work type mapping
        contract_type_raw = clean(row.get("Contract Type")) or "SOW"
        ct_map = {"PO": "SOW", "Renewal": "SOW", "Change Order": "SOW"}
        contract_type = ct_map.get(contract_type_raw, contract_type_raw)
        if contract_type not in ("SOW", "Amendment", "Change Order", "MSA", "NDA"):
            contract_type = "SOW"

        work_type = map_work_type_sow(row.get("Work Type"))
        status = map_sow_status(row.get("Status"))
        rag = rag_clean(row.get("R.A.G. Status"))

        # Budget fields
        monthly_budget = clean_decimal(row.get("Monthly Budget"))
        total_budget = clean_decimal(row.get("Total Budget"))
        block_of_hours = clean_decimal(row.get("Block Of Hours"))

        # Requires PO
        req_po = "Yes" if map_requires_po(row.get("Requires PO")) else "No"  # text: Yes/No
        po_number = (clean(row.get("PO #")) or "")[:100] or None

        # Billing cycle — normalize typos
        cycle_raw = clean(row.get("Cycle"))
        cycle_map = {"quarerly": "Quarterly", "biweekly": "Biweekly", "-": None}
        billing_cycle = cycle_map.get((cycle_raw or "").lower(), cycle_raw)

        # Signed copy
        sc = clean(row.get("Signed Copy"))
        signed_copy = "Yes" if sc == "Yes" else ("No" if sc == "No" else "Pending")  # text

        exec_method = "DocuSign" if clean_bool(row.get("DocuSign")) else None

        # Duration
        dur = clean(row.get("Duration (Months)"))
        # Numeric durations like 60/week → keep as string, pure numbers keep as-is
        if dur and re.match(r"^\d+/\w+$", dur):
            dur = dur.split("/")[0]  # take the number part

        # Project name — required field
        project_name = (clean(row.get("Contract Name/Document Link")) or
                        clean(row.get("Discription of Project")) or
                        f"{client_name} SOW")[:500]

        # Renewal required — Managed Services/Retainer = Yes by business rule
        work_raw = clean(row.get("Work Type")) or ""
        renewal_required = "Yes" if work_raw.lower() in (
            "managed services", "retainer", "retainer services", "remote support"
        ) else "No"  # text: Yes/No/TBD

        # Extract document_name and document_path
        document_name = (clean(row.get("Contract Name/Document Link")) or "")[:500] or None
        document_path = None
        SHAREPOINT_BASE_URL = "https://viscositynorthamerica.sharepoint.com/sites/ProcessMgmt-SharedTeam/Shared Documents/General/CTA/Contracts-  Collateral/"
        
        if wb and doc_link_col_idx:
            try:
                # read_sheet structure: idx 0 = Excel row 3 (row 1=description, row 2=headers, row 3+ data)
                excel_row = idx + 3
                cell = ws.cell(row=excel_row, column=doc_link_col_idx)
                if cell.hyperlink:
                    hyperlink = cell.hyperlink.target
                    # Prepend SharePoint base URL
                    document_path = SHAREPOINT_BASE_URL + hyperlink
                    # log.debug(f"  Row {idx}: Excel row {excel_row} - Found hyperlink: {document_path}")
                else:
                    log.debug(f"  Row {idx}: Excel row {excel_row} - No hyperlink found")
            except Exception as e:
                log.debug(f"  Row {idx}: Error extracting hyperlink: {e}")

        # Find parent agreement (prefer MSA, fall back to any agreement for this client)
        parent_agreement_id = None
        try:
            # First, try to find an MSA for this client
            cur.execute("""
                SELECT agreements_id FROM agreements 
                WHERE client_id = %s AND contract_type = 'MSA'
                ORDER BY effective_date DESC NULLS LAST
                LIMIT 1
            """, (client_id,))
            result = cur.fetchone()
            if result:
                parent_agreement_id = result[0]
            else:
                # Fall back to any agreement for this client
                cur.execute("""
                    SELECT agreements_id FROM agreements 
                    WHERE client_id = %s
                    ORDER BY effective_date DESC NULLS LAST
                    LIMIT 1
                """, (client_id,))
                result = cur.fetchone()
                if result:
                    parent_agreement_id = result[0]
        except Exception as e:
            log.debug(f"  Could not find parent agreement for SOW at row {idx}: {e}")

        # Check if SOW already exists (by client_id + project_name + effective_date)
        eff_date = clean_date(row.get("Effective Date"))
        
        # Build dynamic SQL to handle NULL values
        if project_name and eff_date:
            cur.execute("""
                SELECT sows_id FROM sows 
                WHERE client_id = %s AND project_name = %s AND effective_date = %s
            """, (client_id, project_name, eff_date))
        elif project_name:
            cur.execute("""
                SELECT sows_id FROM sows 
                WHERE client_id = %s AND project_name = %s AND effective_date IS NULL
            """, (client_id, project_name))
        elif eff_date:
            cur.execute("""
                SELECT sows_id FROM sows 
                WHERE client_id = %s AND project_name IS NULL AND effective_date = %s
            """, (client_id, eff_date))
        else:
            cur.execute("""
                SELECT sows_id FROM sows 
                WHERE client_id = %s AND project_name IS NULL AND effective_date IS NULL
            """, (client_id,))
        
        if cur.fetchone():
            # SOW already exists, skip it
            skipped += 1
            log.debug(f"  ⊘ SOW for client_id={client_id}, project='{project_name}' already exists - skipped")
            continue

        cur.execute("""
            INSERT INTO sows (
                client_id, business_unit_id,
                line_of_business, contract_type, work_type,
                project_name, project_contact,
                document_name, document_path, document_stored, signed_copy, execution_method,
                workflow_status,
                date_signed, effective_date, duration_months, expiration_date,
                rag_status,
                parent_agreement_id,
                requires_po, po_number,
                billing_cycle, payment_terms,
                monthly_budget, total_budget, block_of_hours,
                renewal_required, in_subscription_tracker, auto_renewal, rate_reviewed, new_sow_required_on_renewal,
                description, status, archived,
                created_by, last_updated_by
            ) VALUES (
                %s,%s,
                %s,%s,%s,
                %s,%s,
                %s,%s,%s,%s,%s,
                %s,
                %s,%s,%s,%s,
                %s,
                %s,
                %s,%s,
                %s,%s,
                %s,%s,%s,
                %s,%s,%s,%s,%s,
                %s,%s,%s,
                %s,%s
            )
            RETURNING sows_id
        """, (
            client_id, bu_id,
            (clean(row.get("Business Line ")) or "")[:100] or None,
            contract_type,
            work_type,
            project_name,
            (clean(row.get("Project Contact")) or "")[:150] or None,
            # document
            document_name,
            (document_path or "")[:500] or None,
            None,  # document_stored: bytea - stored as NULL initially
            signed_copy,  # boolean
            exec_method,
            # workflow — all historical data = Executed (if signed) or Pre-Sales Draft
            "Executed" if signed_copy else "Pre-Sales Draft",
            # dates
            clean_date(row.get("Date Signed")),
            clean_date(row.get("Effective Date")),
            (dur or "")[:20] or None,
            clean_date(row.get("Expiration Date")),
            rag,
            parent_agreement_id,  # ← Added here
            req_po,
            po_number,
            (billing_cycle or "")[:50] or None,
            (clean(row.get("Payment Terms")) or "")[:50] or None,
            monthly_budget,
            total_budget,
            block_of_hours,
            clean_bool(renewal_required),  # renewal_required: now boolean (converted from "Yes"/"No")
            False,  # in_subscription_tracker: boolean, always false initially
            False,  # auto_renewal: new boolean field, default false
            False,  # rate_reviewed: new boolean field, default false
            False,  # new_sow_required_on_renewal: new boolean field, default false
            (clean(row.get("Discription of Project")) or "")[:5000] or None,
            status,
            1 if status in ("Archived", "Expired") else 0,  # archived: smallint
            None, None,  # created_by, last_updated_by: leave empty
        ))
        inserted += 1

    if wb:
        wb.close()

    log.info(f"  ✓ {inserted} SOWs inserted, {skipped} skipped")
    return inserted, skipped


# ─── STEP 7: PARTNERSHIPS ─────────────────────────────────────────────────────

def import_partnerships(cur):
    log.info("Importing partnerships...")
    df = read_sheet("Partnerships - Referrals and Re")
    df = df[df["Company Name"].apply(lambda x: bool(clean(x)) and clean(x) != "Company Name")]

    inserted = 0
    skipped = 0
    
    # Load Excel workbook for direct hyperlink extraction
    try:
        wb = load_workbook(EXCEL_PATH)
        # Print available sheet names for debugging
        # log.info(f"  Available sheets: {wb.sheetnames}")
        
        sheet_name_to_use = None
        for sheet in wb.sheetnames:
            if "Partnership" in sheet:
                sheet_name_to_use = sheet
                break
        
        if not sheet_name_to_use:
            log.warning(f"  ✗ No 'Partnership' sheet found")
            wb.close()
            wb = None
            agreement_col_idx = None
        else:
            # log.info(f"  Using sheet: '{sheet_name_to_use}'")
            ws = wb[sheet_name_to_use]
            
            # First, let's examine the actual structure
            # log.info(f"  Excel row 1 (description): {[cell.value for cell in ws[1]][:5]}...")
            # log.info(f"  Excel row 2 (headers): {[cell.value for cell in ws[2]][:5]}...")
            
            # Find Agreement column index in row 2 (header row)
            agreement_col_idx = None
            headers = [cell.value for cell in ws[2]]
            
            for idx_col, cell in enumerate(ws[2], 1):
                col_name = cell.value
                if col_name and "agreement" in str(col_name).lower():
                    agreement_col_idx = idx_col
                    # log.info(f"  ✓ Found Agreement column at index {agreement_col_idx} (value: '{col_name}')")
                    break
            
            if not agreement_col_idx:
                log.warning(f"  ✗ Agreement column NOT found in row 2. Headers: {headers}")
    except Exception as e:
        log.warning(f"  Could not load openpyxl workbook for hyperlinks: {e}")
        import traceback
        traceback.print_exc()
        wb = None
        agreement_col_idx = None
    
    for idx, row in df.iterrows():
        company = clean(row.get("Company Name"))
        if not company:
            continue

        ptype_raw = clean(row.get("Referral or Resell")) or ""
        ptype_map = {"referral": "Referral", "resell": "Resell"}
        partner_type = ptype_map.get(ptype_raw.lower())

        agreement_name = clean(row.get("Agreement"))
        document_name = agreement_name  # Use same value for document_name
        document_path = None
        SHAREPOINT_BASE_URL = "https://viscositynorthamerica.sharepoint.com/sites/ProcessMgmt-SharedTeam/Shared Documents/General/CTA/Contracts-  Collateral/"
        
        # Extract hyperlink from Agreement column
        if wb and agreement_col_idx:
            try:
                # read_sheet structure: header=0 reads row 1 as header
                # df.columns = df.iloc[0] uses row 2 as column names
                # df.iloc[1:] removes row 2, so idx 0 = Excel row 3
                excel_row = idx + 3  # Pandas idx 0 = Excel row 3 (row 1=description, row 2=headers, row 3+ data)
                
                cell = ws.cell(row=excel_row, column=agreement_col_idx)
                cell_value = cell.value
                cell_hyperlink = cell.hyperlink
                
                # log.debug(f"  Row {idx}: Excel row {excel_row}, col {agreement_col_idx} -> value='{cell_value}', hyperlink={cell_hyperlink}")
                
                if cell_hyperlink:
                    hyperlink = cell_hyperlink.target
                    # Prepend SharePoint base URL
                    document_path = SHAREPOINT_BASE_URL + hyperlink
                    # log.info(f"    ✓ Found hyperlink: {document_path}")
                else:
                    log.debug(f"    No hyperlink in this cell")
            except Exception as e:
                log.warning(f"  Row {idx}: Error extracting hyperlink: {e}")
                import traceback
                traceback.print_exc()
        
        agr_status_raw = agreement_name or ""
        # Determine status from agreement column content
        if "Pending" in agr_status_raw:
            agr_status = "Pending"
        elif "Redlines" in agr_status_raw or "Redlined" in agr_status_raw:
            agr_status = "Redlines"
        elif "executed" in agr_status_raw.lower() or "final" in agr_status_raw.lower():
            agr_status = "Active"
        else:
            agr_status = "Pending"

        # Check if partnership already exists (by company_name)
        cur.execute("SELECT partnerships_id FROM partnerships WHERE company_name = %s", (company,))
        if cur.fetchone():
            # Partnership already exists, skip it
            skipped += 1
            log.debug(f"  ⊘ Partnership '{company}' already exists - skipped")
            continue

        cur.execute("""
            INSERT INTO partnerships
                (company_name, partner_type,
                 initial_contact_date, last_engagement_date,
                 agreement_status, agreement_name, document_name, document_path, notes)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            company[:255],
            partner_type,
            clean_date(row.get("Initial Contact Date")),
            clean_date(row.get("Last Engagement Date")),
            agr_status,
            (agreement_name or "")[:500] or None,
            (document_name or "")[:500] or None,
            (document_path or "")[:500] or None,
            (clean(row.get("Notes")) or "")[:2000] or None,
        ))
        inserted += 1

    if wb:
        wb.close()

    log.info(f"  ✓ {inserted} partnerships inserted, {skipped} skipped")
    return inserted, skipped


# ─── VALIDATION REPORT ───────────────────────────────────────────────────────

def print_validation_report(cur):
    log.info("\n" + "="*60)
    log.info("POST-IMPORT VALIDATION REPORT")
    log.info("="*60)

    tables = [
        "business_units", "user", "clients", "agreements",
        "agreement_insurance", "sows", "partnerships",
    ]
    for t in tables:
        cur.execute(f"SELECT COUNT(*) FROM {t}")
        count = cur.fetchone()[0]
        log.info(f"  {t:<30} {count:>6} rows")

    log.info("\n--- Data Quality Checks ---")

    # Agreements without clients
    cur.execute("""
        SELECT COUNT(*) FROM agreements a
        LEFT JOIN clients c ON a.client_id = c.client_id
        WHERE c.client_id IS NULL
    """)
    log.info(f"  Agreements with missing client FK:  {cur.fetchone()[0]}")

    # SOWs without clients
    cur.execute("""
        SELECT COUNT(*) FROM sows s
        LEFT JOIN clients c ON s.client_id = c.client_id
        WHERE c.client_id IS NULL
    """)
    log.info(f"  SOWs with missing client FK:        {cur.fetchone()[0]}")

    # Agreements by RAG
    cur.execute("SELECT rag_status, COUNT(*) FROM agreements GROUP BY rag_status ORDER BY 1")
    log.info("\n  Agreements by RAG:")
    for r in cur.fetchall():
        log.info(f"    {r[0]}: {r[1]}")

    # SOWs by status
    cur.execute("SELECT status, COUNT(*) FROM sows GROUP BY status ORDER BY 1")
    log.info("\n  SOWs by status:")
    for r in cur.fetchall():
        log.info(f"    {r[0]}: {r[1]}")

    # SOWs by work_type
    cur.execute("SELECT work_type, COUNT(*) FROM sows GROUP BY work_type ORDER BY 2 DESC")
    log.info("\n  SOWs by work_type:")
    for r in cur.fetchall():
        log.info(f"    {r[0]}: {r[1]}")

    # Clients with no agreements or SOWs
    cur.execute("""
        SELECT COUNT(*) FROM clients c
        WHERE NOT EXISTS (SELECT 1 FROM agreements WHERE client_id = c.client_id)
          AND NOT EXISTS (SELECT 1 FROM sows WHERE client_id = c.client_id)
    """)
    log.info(f"\n  Clients with no agreements/SOWs:    {cur.fetchone()[0]}")

    log.info("="*60)


# ─── DRY RUN MODE ────────────────────────────────────────────────────────────

def dry_run():
    """Parse Excel and print summary without touching DB."""
    log.info("=== DRY RUN MODE (no DB writes) ===")

    qb = pd.read_excel(EXCEL_PATH, sheet_name="QB Consolidated Contacts")
    agr = read_sheet("Agreements")
    sow = read_sheet("Statements of Work")
    par = read_sheet("Partnerships - Referrals and Re")

    agr_clean = agr[agr["Client"].apply(lambda x: bool(clean(x)) and clean(x) != "Client")]
    sow_clean = sow[sow["Client"].apply(lambda x: bool(clean(x)) and clean(x) != "Client")]
    par_clean = par[par["Company Name"].apply(lambda x: bool(clean(x)) and clean(x) != "Company Name")]

    all_clients = (
        set(qb["Customer"].dropna().apply(clean).dropna()) |
        set(agr_clean["Client"].dropna().apply(clean).dropna()) |
        set(sow_clean["Client"].dropna().apply(clean).dropna())
    )

    print(f"\n{'='*60}")
    print(f"  QB Contacts rows:         {len(qb)}")
    print(f"  Agreements rows (data):   {len(agr_clean)}")
    print(f"  SOW rows (data):          {len(sow_clean)}")
    print(f"  Partnership rows (data):  {len(par_clean)}")
    print(f"  Unique clients total:     {len(all_clients)}")
    print(f"\n  Sample clients:")
    for c in sorted(all_clients)[:10]:
        print(f"    - {c}")
    print(f"{'='*60}\n")

    # Show any data quality issues
    print("DATA QUALITY ISSUES:")
    # Agreements with no client name
    blank_clients_agr = agr_clean[agr_clean["Client"].apply(lambda x: not clean(x))]
    print(f"  Agreements with blank client: {len(blank_clients_agr)}")

    # SOWs with no project name or doc link
    no_proj = sow_clean[
        sow_clean["Contract Name/Document Link"].apply(lambda x: not clean(x)) &
        sow_clean["Discription of Project"].apply(lambda x: not clean(x))
    ]
    print(f"  SOWs with no project name:    {len(no_proj)}")

    # Expiration date = '00:00:00' (corrupt)
    corrupt_dates = agr_clean[
        agr_clean["Expiration Date"].apply(lambda x: str(x).strip() == "00:00:00")
    ]
    print(f"  Agreements with corrupt expiry date: {len(corrupt_dates)}")

    print("\nDry run complete. Run without --dry-run to import.")


# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="ETL: Excel → PostgreSQL")
    parser.add_argument("--dry-run", action="store_true", help="Parse only, no DB write")
    parser.add_argument("--sheet", choices=["all","clients","agreements","sows","partnerships"],
                        default="all", help="Which step to run")
    args = parser.parse_args()

    if args.dry_run:
        dry_run()
        return

    log.info("Connecting to PostgreSQL...")
    try:
        conn = get_conn()
    except Exception as e:
        log.error(f"DB connection failed: {e}")
        log.error("Set env vars: DB_HOST, DB_PORT, DB_NAME, DB_USER, DB_PASSWORD")
        sys.exit(1)

    conn.autocommit = False
    cur = conn.cursor()

    try:
        step = args.sheet

        if step in ("all",):
            seed_business_units(cur)
            # seed_system_user(cur)  # Disabled: no user data in Excel

        client_map: dict[str, int] = {}

        if step in ("all", "clients"):
            client_map, _, _ = import_clients(cur)
            conn.commit()

        if step in ("all", "agreements"):
            if not client_map:
                # Re-load from DB if running step independently
                cur.execute("SELECT company_name, client_id FROM clients")
                client_map = {r[0]: r[1] for r in cur.fetchall()}
            _, _, _ = import_agreements(cur, client_map)
            conn.commit()

        if step in ("all", "sows"):
            if not client_map:
                cur.execute("SELECT company_name, client_id FROM clients")
                client_map = {r[0]: r[1] for r in cur.fetchall()}
            _, _ = import_sows(cur, client_map)
            conn.commit()

        if step in ("all", "partnerships"):
            _, _ = import_partnerships(cur)
            conn.commit()

        if step == "all":
            print_validation_report(cur)

        conn.commit()
        log.info("\n✅ All done — transaction committed.")

    except Exception as e:
        conn.rollback()
        log.error(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        cur.close()
        conn.close()


if __name__ == "__main__":
    main()
